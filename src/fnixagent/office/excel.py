"""Excel Expert(P2-9)。

Excel 工作簿创建/读取/公式/数据透视/图表/合并/条件格式/CSV 转换。

专家职责:
  - 创建/读取/编辑 .xlsx 工作簿(多 sheet/公式/图表/条件格式)
  - 数据透视表(基于 pandas)、CSV 互转、多文件合并

底层依赖:
  - openpyxl(读写 .xlsx,可选)
  - pandas(数据透视,可选)

降级策略:
  - 依赖缺失 → ExpertError 提示安装
  - 大文件读用 read_only 模式,IO 后及时 close()
  - 路径穿越/扩展名/大小限制 → _validate_path 拦截
"""
from __future__ import annotations

from typing import Any, Optional

from fnixagent.office.base import BaseExpert, ExpertError, ExpertResult


class ExcelExpert(BaseExpert):
    """Excel 工作簿专家。

    全部方法返回 ExpertResult。

    能力边界:
      - 仅处理 .xlsx(openpyxl 不支持 .xls 旧格式,需 LibreOffice 转换)
      - 公式写入后不计算,需 Excel 打开重算
      - 数据透视表依赖 pandas,大数据集内存受限
    """

    @property
    def name(self) -> str:
        return "excel"

    # ------------------------------------------------------------------
    # 创建与读取
    # ------------------------------------------------------------------

    def create(
        self,
        output_path: str,
        sheets: Optional[list[dict]] = None,
        sheet_name: str = "Sheet1",
        data: Optional[list[list[Any]]] = None,
    ) -> ExpertResult:
        """创建 Excel 工作簿。

        Args:
            output_path: 输出 .xlsx 路径
            sheets: 多 sheet 定义,每项 {name, data:[[row]], headers:[...]}
            sheet_name: 单 sheet 模式下的 sheet 名
            data: 单 sheet 模式下的二维数据(首行可选作 header)

        Returns:
            ExpertResult(output=output_path)
        """
        err = self._validate_path(output_path, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import Workbook
        except ExpertError as e:
            return self._failure(str(e))

        wb = None
        try:
            wb = Workbook()
            # 移除默认 sheet
            default_ws = wb.active
            wb.remove(default_ws)

            if sheets:
                for spec in sheets:
                    ws = wb.create_sheet(title=spec.get("name", f"Sheet{len(wb.sheetnames)+1}"))
                    headers = spec.get("headers")
                    rows = spec.get("data", [])
                    if headers:
                        ws.append(headers)
                    for row in rows:
                        ws.append(row)
            else:
                ws = wb.create_sheet(title=sheet_name)
                if data:
                    for row in data:
                        ws.append(row)
            # 至少保留一个 sheet
            if not wb.sheetnames:
                wb.create_sheet(title="Sheet1")
            wb.save(output_path)
            return self._success(output_path, sheets=len(wb.sheetnames))
        except (PermissionError, IOError) as e:
            return self._failure(f"create excel IO failed: {e}")
        except Exception as e:
            return self._failure(f"create excel failed: {e}")
        finally:
            # 写模式工作簿也建议关闭,释放 zip 文件句柄
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def read(
        self,
        path: str,
        sheet_name: Optional[str] = None,
        max_rows: Optional[int] = None,
        with_header: bool = True,
    ) -> ExpertResult:
        """读取 Excel 内容为二维列表。

        Args:
            path: .xlsx 文件路径
            sheet_name: 指定 sheet;None 取第一个
            max_rows: 限制最大行数
            with_header: 是否将首行作为 headers 返回

        Returns:
            ExpertResult(output={headers, rows, sheet_names})
        """
        err = self._validate_path(
            path, must_exist=True, allowed_exts=("xlsx",)
        )
        if err:
            return self._failure(err)
        if max_rows is not None:
            err = self._validate_int(max_rows, "max_rows", min_value=0)
            if err:
                return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import load_workbook
        except ExpertError as e:
            return self._failure(str(e))

        wb = None
        try:
            # read_only 模式:流式读取,适合大文件
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            ws = wb[sheet_name] if sheet_name else wb[sheet_names[0]]
            rows_iter = ws.iter_rows(values_only=True)
            all_rows = list(rows_iter)
            if max_rows is not None:
                all_rows = all_rows[:max_rows]
            headers = list(all_rows[0]) if all_rows and with_header else []
            data_rows = all_rows[1:] if with_header else all_rows
            return self._success(
                output={"headers": headers, "rows": data_rows, "sheet_names": sheet_names},
                total_rows=len(data_rows),
            )
        except KeyError as e:
            return self._failure(f"sheet not found: {e}")
        except (PermissionError, IOError) as e:
            return self._failure(f"read excel IO failed: {e}")
        except Exception as e:
            return self._failure(f"read excel failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    # ------------------------------------------------------------------
    # 公式与计算
    # ------------------------------------------------------------------

    def formula(
        self,
        path: str,
        cell: str,
        formula: str,
        sheet_name: Optional[str] = None,
    ) -> ExpertResult:
        """向指定单元格写入公式。

        Args:
            path: .xlsx 文件路径(原地修改)
            cell: 单元格地址(如 "B10")
            formula: 公式字符串(如 "=SUM(B1:B9)")
            sheet_name: 指定 sheet;None 取活动 sheet

        Returns:
            ExpertResult(output=cell)
        """
        err = self._validate_path(
            path, must_exist=True, allowed_exts=("xlsx",)
        )
        if err:
            return self._failure(err)
        err = self._validate_string(cell, "cell")
        if err:
            return self._failure(err)
        err = self._validate_string(formula, "formula")
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import load_workbook
        except ExpertError as e:
            return self._failure(str(e))

        wb = None
        try:
            wb = load_workbook(path)
            ws = wb[sheet_name] if sheet_name else wb.active
            ws[cell] = formula
            wb.save(path)
            return self._success(cell, formula=formula, sheet=ws.title)
        except KeyError as e:
            return self._failure(f"sheet not found: {e}")
        except (PermissionError, IOError) as e:
            return self._failure(f"set formula IO failed: {e}")
        except Exception as e:
            return self._failure(f"set formula failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    # ------------------------------------------------------------------
    # 数据透视
    # ------------------------------------------------------------------

    def pivot_table(
        self,
        path: str,
        output_path: str,
        source_sheet: Optional[str],
        rows: list[str],
        values: list[str],
        cols: Optional[list[str]] = None,
        agg_func: str = "sum",
    ) -> ExpertResult:
        """生成数据透视表(基于 pandas)。

        Args:
            path: 源 .xlsx 路径
            output_path: 输出 .xlsx 路径
            source_sheet: 源 sheet;None 取第一个
            rows: 行分组字段
            values: 聚合值字段
            cols: 列分组字段(可选)
            agg_func: 聚合函数 sum/mean/count/max/min

        Returns:
            ExpertResult(output=output_path)
        """
        err = self._validate_path(
            path, must_exist=True, allowed_exts=("xlsx",)
        )
        if err:
            return self._failure(err)
        err = self._validate_path(output_path, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)
        if not rows or not values:
            return self._failure("rows and values must be non-empty")

        try:
            self._require_lib("pandas")
            import pandas as pd
        except ExpertError as e:
            return self._failure(str(e))

        try:
            df = pd.read_excel(path, sheet_name=source_sheet or 0)
            pt = pd.pivot_table(
                df, index=rows, values=values, columns=cols, aggfunc=agg_func, fill_value=0
            )
            pt.to_excel(output_path)
            return self._success(
                output_path,
                rows=len(pt),
                cols=len(pt.columns),
                agg_func=agg_func,
            )
        except (PermissionError, IOError) as e:
            return self._failure(f"pivot_table IO failed: {e}")
        except Exception as e:
            return self._failure(f"pivot_table failed: {e}")

    # ------------------------------------------------------------------
    # 图表
    # ------------------------------------------------------------------

    def chart(
        self,
        path: str,
        sheet_name: Optional[str],
        chart_type: str,
        data_range: str,
        title: str = "",
        anchor: str = "E1",
    ) -> ExpertResult:
        """在 sheet 中插入图表。

        Args:
            path: .xlsx 路径(原地修改)
            sheet_name: 目标 sheet;None 取活动 sheet
            chart_type: bar / line / pie / scatter
            data_range: 数据范围(如 "A1:B10")
            title: 图表标题
            anchor: 图表锚点单元格

        Returns:
            ExpertResult(output=anchor)
        """
        err = self._validate_path(
            path, must_exist=True, allowed_exts=("xlsx",)
        )
        if err:
            return self._failure(err)
        err = self._validate_string(data_range, "data_range")
        if err:
            return self._failure(err)
        err = self._validate_string(chart_type, "chart_type")
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import load_workbook
            from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
        except ExpertError as e:
            return self._failure(str(e))

        wb = None
        try:
            wb = load_workbook(path)
            ws = wb[sheet_name] if sheet_name else wb.active
            # 解析 data_range,如 "A1:B10" → min_col=1, max_col=2, min_row=1, max_row=10
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(data_range)
            data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

            ct = chart_type.lower()
            if ct == "bar":
                chart = BarChart()
            elif ct == "line":
                chart = LineChart()
            elif ct == "pie":
                chart = PieChart()
            elif ct == "scatter":
                chart = ScatterChart()
            else:
                return self._failure(f"unsupported chart_type: {chart_type}")
            chart.add_data(data_ref, titles_from_data=True)
            if title:
                chart.title = title
            ws.add_chart(chart, anchor)
            wb.save(path)
            return self._success(anchor, chart_type=ct)
        except KeyError as e:
            return self._failure(f"sheet not found: {e}")
        except (PermissionError, IOError) as e:
            return self._failure(f"insert chart IO failed: {e}")
        except Exception as e:
            return self._failure(f"insert chart failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    # ------------------------------------------------------------------
    # 合并
    # ------------------------------------------------------------------

    def merge(
        self,
        paths: list[str],
        output_path: str,
        sheet_names: Optional[list[str]] = None,
    ) -> ExpertResult:
        """合并多个 Excel 文件为一个多 sheet 工作簿。

        Args:
            paths: 源 .xlsx 路径列表
            output_path: 输出 .xlsx 路径
            sheet_names: 每个 sheet 的命名;None 用源文件名

        Returns:
            ExpertResult(output=output_path, sheets=N)
        """
        if not paths:
            return self._failure("paths is empty")
        for p in paths:
            err = self._validate_path(
                p, must_exist=True, allowed_exts=("xlsx",)
            )
            if err:
                return self._failure(err)
        err = self._validate_path(output_path, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import Workbook, load_workbook
        except ExpertError as e:
            return self._failure(str(e))

        merged = None
        src_wb = None
        try:
            merged = Workbook()
            merged.remove(merged.active)
            for idx, p in enumerate(paths):
                src_wb = load_workbook(p, read_only=True)
                src_ws = src_wb.active
                name = (sheet_names[idx] if sheet_names and idx < len(sheet_names)
                        else f"Sheet{idx+1}")
                # sheet 名最长 31 字符(Excel 限制)
                dst_ws = merged.create_sheet(title=name[:31])
                for row in src_ws.iter_rows(values_only=True):
                    dst_ws.append(row)
                src_wb.close()
                src_wb = None
            if not merged.sheetnames:
                merged.create_sheet(title="Sheet1")
            merged.save(output_path)
            return self._success(output_path, sheets=len(merged.sheetnames))
        except (PermissionError, IOError) as e:
            return self._failure(f"merge excel IO failed: {e}")
        except Exception as e:
            return self._failure(f"merge excel failed: {e}")
        finally:
            # 确保源工作簿与目标工作簿都释放
            if src_wb is not None:
                try:
                    src_wb.close()
                except Exception:
                    pass
            if merged is not None:
                try:
                    merged.close()
                except Exception:
                    pass

    # ------------------------------------------------------------------
    # 条件格式
    # ------------------------------------------------------------------

    def conditional_format(
        self,
        path: str,
        cell_range: str,
        rule_type: str,
        sheet_name: Optional[str] = None,
        value: Any = None,
        color: str = "FFC7CE",
    ) -> ExpertResult:
        """添加条件格式规则。

        Args:
            path: .xlsx 路径(原地修改)
            cell_range: 应用范围(如 "B2:B100")
            rule_type: cell / color_scale / data_bar / top
            sheet_name: 目标 sheet;None 取活动 sheet
            value: 规则阈值(cell 类型用比较值)
            color: 高亮颜色(默认浅红)

        Returns:
            ExpertResult(output=cell_range)
        """
        err = self._validate_path(
            path, must_exist=True, allowed_exts=("xlsx",)
        )
        if err:
            return self._failure(err)
        err = self._validate_string(cell_range, "cell_range")
        if err:
            return self._failure(err)
        err = self._validate_string(rule_type, "rule_type")
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import load_workbook
            from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, Rule
            from openpyxl.styles import PatternFill
        except ExpertError as e:
            return self._failure(str(e))

        wb = None
        try:
            wb = load_workbook(path)
            ws = wb[sheet_name] if sheet_name else wb.active
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            rt = rule_type.lower()
            if rt == "cell":
                # value 形如 ">80"
                if not value:
                    return self._failure("cell rule requires value, e.g. '>80'")
                op = "greaterThan"
                if value.startswith(">="):
                    op, val = "greaterThanOrEqual", value[2:]
                elif value.startswith(">"):
                    op, val = "greaterThan", value[1:]
                elif value.startswith("<="):
                    op, val = "lessThanOrEqual", value[2:]
                elif value.startswith("<"):
                    op, val = "lessThan", value[1:]
                elif value.startswith("==") or value.startswith("="):
                    op, val = "equal", value.lstrip("=")
                else:
                    val = value
                try:
                    val_num = float(val)
                except ValueError:
                    val_num = val
                rule = CellIsRule(operator=op, formula=[val_num], fill=fill)
            elif rt == "color_scale":
                rule = ColorScaleRule(
                    start_type="min", start_color="63BE7B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="F8696B",
                )
            elif rt == "data_bar":
                rule = DataBarRule(start_type="min", end_type="max", color=color)
            elif rt == "top":
                rule = Rule(type="top10", rank=int(value or 10), percent=False, bottom=False, dxf=None)
            else:
                return self._failure(f"unsupported rule_type: {rule_type}")
            ws.conditional_formatting.add(cell_range, rule)
            wb.save(path)
            return self._success(cell_range, rule_type=rt)
        except KeyError as e:
            return self._failure(f"sheet not found: {e}")
        except (PermissionError, IOError) as e:
            return self._failure(f"conditional_format IO failed: {e}")
        except Exception as e:
            return self._failure(f"conditional_format failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    # ------------------------------------------------------------------
    # CSV 转换
    # ------------------------------------------------------------------

    def to_csv(
        self,
        path: str,
        output_path: str,
        sheet_name: Optional[str] = None,
        encoding: str = "utf-8-sig",
    ) -> ExpertResult:
        """将 Excel sheet 导出为 CSV。

        Args:
            path: 源 .xlsx 路径
            output_path: 输出 .csv 路径
            sheet_name: 指定 sheet;None 取第一个
            encoding: CSV 编码(默认 utf-8-sig 兼容 Excel 打开)

        Returns:
            ExpertResult(output=output_path)
        """
        err = self._validate_path(
            path, must_exist=True, allowed_exts=("xlsx",)
        )
        if err:
            return self._failure(err)
        err = self._validate_path(output_path, allowed_exts=("csv",), max_size=None)
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            import csv
            from openpyxl import load_workbook
        except ExpertError as e:
            return self._failure(str(e))

        wb = None
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
            row_count = 0
            # newline="" 避免 Windows 多余空行;utf-8-sig 兼容 Excel 中文
            with open(output_path, "w", newline="", encoding=encoding) as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
                    row_count += 1
            return self._success(output_path, rows=row_count, encoding=encoding)
        except KeyError as e:
            return self._failure(f"sheet not found: {e}")
        except (PermissionError, IOError) as e:
            return self._failure(f"to_csv IO failed: {e}")
        except Exception as e:
            return self._failure(f"to_csv failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
