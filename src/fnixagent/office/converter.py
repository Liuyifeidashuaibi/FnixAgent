"""Converter Expert(P2-9)。

文档格式转换:docx/xlsx/pptx/pdf/html/markdown/csv 互转。

专家职责:
  - 直接转换:Python 库处理(xlsx↔csv↔json、md→html/pdf、html→pdf)
  - 兜底转换:LibreOffice headless 处理(docx/xlsx/pptx 互转)

底层依赖:
  - openpyxl(xlsx/csv)、markdown(md→html)、pdfkit(html→pdf,需 wkhtmltopdf)
  - LibreOffice(兜底,需安装 soffice)

降级策略:
  - 直接转换器依赖缺失 → 转 LibreOffice 兜底
  - LibreOffice 不可用 → 返回 _failure 提示安装
  - subprocess 超时/CalledProcessError 单独捕获,给出可读错误
  - 临时 HTML 文件用 tempfile.NamedTemporaryFile + finally 清理
"""
from __future__ import annotations

import contextlib
import os
import shutil
import subprocess
import tempfile
from typing import Any, Optional

from fnixagent.office.base import BaseExpert, ExpertError, ExpertResult


# 支持的转换对(直接由 Python 库处理)
_DIRECT_CONVERSIONS: dict[tuple[str, str], str] = {
    ("xlsx", "csv"): "excel_to_csv",
    ("csv", "xlsx"): "csv_to_excel",
    ("csv", "json"): "csv_to_json",
    ("json", "csv"): "json_to_csv",
    ("json", "xlsx"): "json_to_excel",
    ("html", "pdf"): "html_to_pdf",
    ("md", "html"): "md_to_html",
    ("md", "pdf"): "md_to_pdf",
}

# LibreOffice 兜底的转换对
_LIBREOFFICE_CONVERSIONS = {
    ("docx", "pdf"), ("docx", "html"), ("docx", "txt"),
    ("xlsx", "pdf"), ("xlsx", "html"), ("xlsx", "csv"),
    ("pptx", "pdf"), ("pptx", "html"), ("pptx", "png"),
    ("doc", "pdf"), ("doc", "docx"), ("xls", "xlsx"),
    ("ppt", "pptx"), ("ppt", "pdf"),
    ("odt", "docx"), ("ods", "xlsx"), ("odp", "pptx"),
}


class ConverterExpert(BaseExpert):
    """文档格式转换专家。

    全部方法返回 ExpertResult。

    能力边界:
      - 直接转换覆盖 xlsx/csv/json/md/html 互转
      - docx/xlsx/pptx 互转依赖 LibreOffice
      - 不支持加密文档(需先解密)
    """

    @property
    def name(self) -> str:
        return "converter"

    # ------------------------------------------------------------------
    # 主入口
    # ------------------------------------------------------------------

    def convert(
        self,
        source_path: str,
        output_path: str,
        target_format: Optional[str] = None,
        **options: Any,
    ) -> ExpertResult:
        """单文件格式转换。

        Args:
            source_path: 源文件路径
            output_path: 输出文件路径
            target_format: 目标格式(扩展名,如 "pdf");None 从 output_path 推断
            **options: 转换选项(透传给具体转换器)

        Returns:
            ExpertResult(output=output_path)
        """
        # 源文件校验:必须存在
        err = self._validate_path(source_path, must_exist=True)
        if err:
            return self._failure(err)
        if not output_path or not isinstance(output_path, str):
            return self._failure("output_path must be a non-empty string")

        src_ext = self._ext(source_path)
        tgt_ext = target_format.lower().lstrip(".") if target_format else self._ext(output_path)
        if not tgt_ext:
            return self._failure("cannot determine target format")

        # 1. 直接转换
        handler_name = _DIRECT_CONVERSIONS.get((src_ext, tgt_ext))
        if handler_name:
            handler = getattr(self, handler_name, None)
            if handler:
                return handler(source_path, output_path, **options)

        # 2. 相同格式(复制)
        if src_ext == tgt_ext:
            try:
                shutil.copy2(source_path, output_path)
            except (PermissionError, IOError) as e:
                return self._failure(f"copy IO failed: {e}")
            return self._success(output_path, mode="copy")

        # 3. 原生 Word COM 转换(docx→pdf,色彩保真度最高)
        #    仅 Windows + 已安装 MS Word 时可用,降级到 LibreOffice
        if (src_ext, tgt_ext) == ("docx", "pdf") and options.get(
            "prefer_word", True
        ):
            result = self._convert_via_word_com(source_path, output_path)
            if result.success:
                return result
            # Word COM 失败,降级到 LibreOffice

        # 4. LibreOffice 兜底
        if (src_ext, tgt_ext) in _LIBREOFFICE_CONVERSIONS:
            return self._convert_via_libreoffice(source_path, output_path, tgt_ext)

        return self._failure(
            f"unsupported conversion: {src_ext} -> {tgt_ext}. "
            f"Supported: {self.supported_conversions()}",
        )

    def batch_convert(
        self,
        source_paths: list[str],
        output_dir: str,
        target_format: str,
        keep_filename: bool = True,
    ) -> ExpertResult:
        """批量格式转换。

        Args:
            source_paths: 源文件路径列表
            output_dir: 输出目录
            target_format: 目标格式(扩展名,如 "pdf")
            keep_filename: True 保留原文件名(改扩展名);False 用序号

        Returns:
            ExpertResult(output={success: [...], failed: [...]})
        """
        if not source_paths:
            return self._failure("source_paths is empty")
        if not output_dir or not isinstance(output_dir, str):
            return self._failure("output_dir must be a non-empty string")
        if not target_format or not isinstance(target_format, str):
            return self._failure("target_format must be a non-empty string")
        try:
            os.makedirs(output_dir, exist_ok=True)
        except (PermissionError, IOError) as e:
            return self._failure(f"create output_dir failed: {e}")
        tgt = target_format.lower().lstrip(".")
        success_list: list[str] = []
        failed_list: list[dict] = []
        for idx, src in enumerate(source_paths):
            base = os.path.splitext(os.path.basename(src))[0] if keep_filename else f"file_{idx+1}"
            out = os.path.join(output_dir, f"{base}.{tgt}")
            r = self.convert(src, out, target_format=tgt)
            if r.success:
                success_list.append(out)
            else:
                failed_list.append({"file": src, "error": r.error})
        return self._success(
            output={"success": success_list, "failed": failed_list},
            total=len(source_paths),
            succeeded=len(success_list),
            failed=len(failed_list),
        )

    def supported_conversions(self) -> dict[str, list[str]]:
        """返回支持的转换映射 {src_ext: [tgt_ext, ...]}。"""
        result: dict[str, list[str]] = {}
        for src, tgt in list(_DIRECT_CONVERSIONS.keys()) | _LIBREOFFICE_CONVERSIONS:
            result.setdefault(src, [])
            if tgt not in result[src]:
                result[src].append(tgt)
        return result

    # ------------------------------------------------------------------
    # 内部转换器
    # ------------------------------------------------------------------

    def excel_to_csv(self, src: str, out: str, **opts: Any) -> ExpertResult:
        err = self._validate_path(src, must_exist=True, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("csv",), max_size=None)
        if err:
            return self._failure(err)
        try:
            self._require_lib("openpyxl")
            import csv
            from openpyxl import load_workbook
        except ExpertError as e:
            return self._failure(str(e))
        wb = None
        try:
            sheet = opts.get("sheet")
            wb = load_workbook(src, read_only=True, data_only=True)
            ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
            with open(out, "w", newline="", encoding=opts.get("encoding", "utf-8-sig")) as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    w.writerow(["" if v is None else v for v in row])
            return self._success(out, mode="excel_to_csv")
        except KeyError as e:
            return self._failure(f"sheet not found: {e}")
        except (PermissionError, IOError) as e:
            return self._failure(f"excel_to_csv IO failed: {e}")
        except Exception as e:
            return self._failure(f"excel_to_csv failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def csv_to_excel(self, src: str, out: str, **opts: Any) -> ExpertResult:
        err = self._validate_path(src, must_exist=True, allowed_exts=("csv",))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)
        try:
            self._require_lib("openpyxl")
            import csv
            from openpyxl import Workbook
        except ExpertError as e:
            return self._failure(str(e))
        wb = None
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = opts.get("sheet_name", "Sheet1")
            # utf-8-sig 兼容 Excel 导出的 CSV
            with open(src, "r", encoding=opts.get("encoding", "utf-8-sig")) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
            wb.save(out)
            return self._success(out, mode="csv_to_excel")
        except (PermissionError, IOError) as e:
            return self._failure(f"csv_to_excel IO failed: {e}")
        except Exception as e:
            return self._failure(f"csv_to_excel failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def csv_to_json(self, src: str, out: str, **opts: Any) -> ExpertResult:
        err = self._validate_path(src, must_exist=True, allowed_exts=("csv",))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("json",), max_size=None)
        if err:
            return self._failure(err)
        import csv
        import json
        try:
            with open(src, "r", encoding=opts.get("encoding", "utf-8-sig")) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            with open(out, "w", encoding="utf-8") as f:
                # ensure_ascii=False 保留中文可读
                json.dump(rows, f, ensure_ascii=False, indent=2)
            return self._success(out, mode="csv_to_json", rows=len(rows))
        except (PermissionError, IOError) as e:
            return self._failure(f"csv_to_json IO failed: {e}")
        except Exception as e:
            return self._failure(f"csv_to_json failed: {e}")

    def json_to_csv(self, src: str, out: str, **opts: Any) -> ExpertResult:
        err = self._validate_path(src, must_exist=True, allowed_exts=("json",))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("csv",), max_size=None)
        if err:
            return self._failure(err)
        import csv
        import json
        try:
            with open(src, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, list) or not data:
                return self._failure("json_to_csv requires a non-empty list of objects")
            headers = list(data[0].keys())
            with open(out, "w", newline="", encoding=opts.get("encoding", "utf-8-sig")) as f:
                w = csv.DictWriter(f, fieldnames=headers)
                w.writeheader()
                w.writerows(data)
            return self._success(out, mode="json_to_csv", rows=len(data))
        except (PermissionError, IOError) as e:
            return self._failure(f"json_to_csv IO failed: {e}")
        except Exception as e:
            return self._failure(f"json_to_csv failed: {e}")

    def json_to_excel(self, src: str, out: str, **opts: Any) -> ExpertResult:
        err = self._validate_path(src, must_exist=True, allowed_exts=("json",))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)
        try:
            self._require_lib("openpyxl")
            import json
            from openpyxl import Workbook
        except ExpertError as e:
            return self._failure(str(e))
        wb = None
        try:
            with open(src, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, list) or not data:
                return self._failure("json_to_excel requires a non-empty list of objects")
            wb = Workbook()
            ws = wb.active
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h) for h in headers])
            wb.save(out)
            return self._success(out, mode="json_to_excel", rows=len(data))
        except (PermissionError, IOError) as e:
            return self._failure(f"json_to_excel IO failed: {e}")
        except Exception as e:
            return self._failure(f"json_to_excel failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def html_to_pdf(self, src: str, out: str, **opts: Any) -> ExpertResult:
        err = self._validate_path(src, must_exist=True, allowed_exts=("html", "htm"))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("pdf",))
        if err:
            return self._failure(err)
        try:
            self._require_lib("pdfkit")
            import pdfkit
        except ExpertError as e:
            # 退到 LibreOffice
            return self._convert_via_libreoffice(src, out, "pdf")
        try:
            options = {"quiet": "", "encoding": "UTF-8"}
            options.update(opts.get("pdfkit_options", {}))
            pdfkit.from_file(src, out, options=options)
            return self._success(out, mode="html_to_pdf")
        except (PermissionError, IOError) as e:
            return self._failure(f"html_to_pdf IO failed: {e}")
        except Exception as e:
            return self._failure(f"html_to_pdf failed: {e}")

    def md_to_html(self, src: str, out: str, **opts: Any) -> ExpertResult:
        err = self._validate_path(src, must_exist=True, allowed_exts=("md",))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("html", "htm"), max_size=None)
        if err:
            return self._failure(err)
        try:
            self._require_lib("markdown")
            import markdown
        except ExpertError as e:
            return self._failure(str(e))
        try:
            with open(src, "r", encoding="utf-8") as f:
                md_text = f.read()
            extensions = opts.get("extensions", ["tables", "fenced_code", "toc"])
            html_body = markdown.markdown(md_text, extensions=extensions)
            title = opts.get("title", os.path.splitext(os.path.basename(src))[0])
            full_html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>{title}</title>
<style>{opts.get("css", "body{font-family:sans-serif;max-width:800px;margin:2em auto;padding:0 1em;}")}</style>
</head>
<body>
{html_body}
</body>
</html>"""
            with open(out, "w", encoding="utf-8") as f:
                f.write(full_html)
            return self._success(out, mode="md_to_html")
        except (PermissionError, IOError) as e:
            return self._failure(f"md_to_html IO failed: {e}")
        except Exception as e:
            return self._failure(f"md_to_html failed: {e}")

    def md_to_pdf(self, src: str, out: str, **opts: Any) -> ExpertResult:
        # 先 md -> html,再 html -> pdf
        err = self._validate_path(src, must_exist=True, allowed_exts=("md",))
        if err:
            return self._failure(err)
        err = self._validate_path(out, allowed_exts=("pdf",))
        if err:
            return self._failure(err)
        # 用 ExitStack 确保临时 HTML 一定被清理,即使中间步骤抛异常
        with contextlib.ExitStack() as stack:
            tmp = tempfile.NamedTemporaryFile(
                mode="w", suffix=".html", delete=False, encoding="utf-8"
            )
            html_path = tmp.name
            tmp.close()
            stack.callback(self._safe_remove_file, html_path)
            r = self.md_to_html(src, html_path, **opts)
            if not r.success:
                return r
            return self.html_to_pdf(html_path, out, **opts)

    @staticmethod
    def _safe_remove_file(path: str) -> None:
        """安全删除临时文件,忽略错误。"""
        try:
            os.remove(path)
        except (OSError, FileNotFoundError):
            pass

    # ------------------------------------------------------------------
    # 原生 Word COM 转换(Windows,色彩保真度最高)
    # ------------------------------------------------------------------

    def _convert_via_word_com(self, src: str, out: str) -> ExpertResult:
        """通过 Windows COM 调用原生 MS Word 进行 docx→pdf 转换。

        优势:
          - 色彩保真度 100%(与 Word 中看到的效果完全一致)
          - 正确处理 CMYK 图片、ICC 配置文件、复杂排版
          - 支持 LibreOffice 无法渲染的高级特性

        降级条件:
          - 非 Windows 平台 → 返回 _failure,上层降级到 LibreOffice
          - 未安装 MS Word → 返回 _failure
          - COM 调用异常 → 返回 _failure

        Args:
            src: 源 .docx 文件路径
            out: 输出 .pdf 文件路径

        Returns:
            ExpertResult(output=out, mode="word_com")
        """
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return self._failure("pywin32 not available, cannot use Word COM")

        word_app = None
        try:
            pythoncom.CoInitialize()
            word_app = win32com.client.DispatchEx("Word.Application")
            word_app.Visible = False
            word_app.DisplayAlerts = False

            abs_src = os.path.abspath(src)
            abs_out = os.path.abspath(out)
            os.makedirs(os.path.dirname(abs_out), exist_ok=True)

            doc = word_app.Documents.Open(abs_src, ReadOnly=True)
            # 17 = wdFormatPDF
            doc.SaveAs(abs_out, FileFormat=17)
            doc.Close(False)

            return self._success(abs_out, mode="word_com")
        except Exception as e:
            return self._failure(f"Word COM conversion failed: {e}")
        finally:
            if word_app is not None:
                try:
                    word_app.Quit()
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    # ------------------------------------------------------------------
    # LibreOffice 兜底
    # ------------------------------------------------------------------

    def _convert_via_libreoffice(
        self, src: str, out: str, target_format: str
    ) -> ExpertResult:
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            # Windows 常见路径
            for candidate in [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            ]:
                if os.path.exists(candidate):
                    soffice = candidate
                    break
        if not soffice:
            return self._failure(
                f"LibreOffice not found. Conversion {self._ext(src)} -> {target_format} "
                f"requires LibreOffice. Install it or use a direct Python converter.",
            )

        try:
            out_dir = os.path.dirname(os.path.abspath(out))
            os.makedirs(out_dir, exist_ok=True)
            subprocess.run(
                [soffice, "--headless", "--convert-to", target_format,
                 "--outdir", out_dir, src],
                check=True, capture_output=True, timeout=180,
            )
            # LibreOffice 输出文件名 = 源文件名 + .target_format
            expected = os.path.join(
                out_dir,
                os.path.splitext(os.path.basename(src))[0] + "." + target_format,
            )
            if expected != os.path.abspath(out):
                if os.path.exists(expected):
                    shutil.move(expected, out)
                else:
                    return self._failure(
                        f"LibreOffice conversion produced no output (expected: {expected})",
                    )
            return self._success(out, mode="libreoffice")
        except subprocess.CalledProcessError as e:
            stderr_msg = e.stderr.decode("utf-8", errors="replace") if e.stderr else str(e)
            return self._failure(f"LibreOffice conversion failed: {stderr_msg}")
        except subprocess.TimeoutExpired:
            return self._failure("LibreOffice conversion timeout (180s)")
        except (PermissionError, IOError) as e:
            return self._failure(f"LibreOffice conversion IO failed: {e}")
        except Exception as e:
            return self._failure(f"LibreOffice conversion failed: {e}")

    # ------------------------------------------------------------------
    # 工具
    # ------------------------------------------------------------------

    @staticmethod
    def _ext(path: str) -> str:
        return os.path.splitext(path)[1].lstrip(".").lower()
