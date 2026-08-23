"""待确认清单导出(Phase 6.3)。

将无法自动确定的答案/内容导出为 pending 清单,供人工补全:
  - LLM 置信度低的答案
  - 题库未命中的题
  - 乱码不可恢复的答案

支持:
  - 导出 xlsx(表头加粗 + 背景色 / 列宽自适应 / 冻结首行)
  - 导出 CSV(utf-8-sig 兼容 Excel 中文)
  - 从 xlsx 导入人工确认的答案(按"题号"+"确认答案"列)
  - 统计(总数/待确认/已确认/平均置信度)

依赖 openpyxl(可选,缺失时返回 ExpertError 提示安装)。
"""

# -*- coding: utf-8 -*-
# Copyright (C) 2026 FnixAgent. All rights reserved.
# Software Name: FnixAgent 智能工作台系统 V1.0
# This software and its source code are proprietary and confidential.
# Unauthorized copying, modification, distribution, or use is strictly prohibited.

from __future__ import annotations

import csv
import logging
import uuid
from dataclasses import dataclass
from typing import Any

from fnixagent.office.base import BaseExpert, ExpertError, ExpertResult

_logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 数据结构
# ---------------------------------------------------------------------------


@dataclass
class PendingItem:
    """待确认项。

    Attributes:
        item_id: 待确认项唯一 ID
        task_id: 所属任务 ID
        question_num: 题号
        stem: 题干(截断到 100 字)
        options: 选项列表(如 ["A. ...", "B. ...", ...])
        garbled_answer: 原始乱码答案
        suggested_answer: 建议答案(来自 LLM)
        confidence: 置信度(0.0~1.0)
        reason: 待确认原因(如"LLM置信度低"/"题库未命中"/"乱码不可恢复")
        status: 状态 pending / confirmed / rejected
        confirmed_answer: 人工确认的答案
    """

    item_id: str
    task_id: str
    question_num: str
    stem: str
    options: list[str]
    garbled_answer: str
    suggested_answer: str | None
    confidence: float
    reason: str
    status: str = "pending"  # pending / confirmed / rejected
    confirmed_answer: str | None = None  # 人工确认的答案


# ---------------------------------------------------------------------------
# 导出列定义
# ---------------------------------------------------------------------------

# 列顺序:(列标题, 对应字段取值方式)
# 取值方式:字符串表示 PendingItem 同名字段;"optN" 表示 options[N]
_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("题号", "question_num"),
    ("题干", "stem"),
    ("选项A", "opt0"),
    ("选项B", "opt1"),
    ("选项C", "opt2"),
    ("选项D", "opt3"),
    ("乱码答案", "garbled_answer"),
    ("建议答案", "suggested_answer"),
    ("置信度", "confidence"),
    ("原因", "reason"),
    ("确认状态", "status"),
    ("确认答案", "confirmed_answer"),
]


def _get_option(options: list[str], idx: int) -> str:
    """安全取选项第 idx 项,越界返回空串。"""
    return options[idx] if idx < len(options) else ""


def _to_cell(value: Any) -> Any:
    """将字段值转为单元格写入值(None → 空串,便于 CSV/Excel 显示)。"""
    if value is None:
        return ""
    return value


# ---------------------------------------------------------------------------
# PendingExporter
# ---------------------------------------------------------------------------


class PendingExporter(BaseExpert):
    """待确认清单导出器。

    用法:
        exp = PendingExporter()
        exp.add_item("t1", "Q1", "题干...", ["A","B","C","D"], "??", "B", 0.3, "乱码不可恢复")
        exp.export_excel("pending.xlsx")
        # 人工在 xlsx 中填写"确认答案"列后
        exp.import_confirmed("pending.xlsx")
    """

    @property
    def name(self) -> str:
        return "pending_exporter"

    def __init__(self) -> None:
        self._items: list[PendingItem] = []

    # ------------------------------------------------------------------
    # 增删查
    # ------------------------------------------------------------------

    def add_item(
        self,
        task_id: str,
        question_num: str,
        stem: str,
        options: list[str],
        garbled_answer: str,
        suggested_answer: str | None,
        confidence: float,
        reason: str,
    ) -> str:
        """添加待确认项,返回 item_id。

        Args:
            task_id: 任务 ID
            question_num: 题号
            stem: 题干(自动截断到 100 字)
            options: 选项列表
            garbled_answer: 原始乱码答案
            suggested_answer: 建议答案(来自 LLM,可空)
            confidence: 置信度(0.0~1.0)
            reason: 待确认原因

        Returns:
            item_id
        """
        item = PendingItem(
            item_id=f"pi-{uuid.uuid4().hex[:12]}",
            task_id=task_id,
            question_num=question_num,
            stem=(stem or "")[:100],  # 题干截断到 100 字
            options=list(options or []),
            garbled_answer=garbled_answer or "",
            suggested_answer=suggested_answer,
            confidence=float(confidence),
            reason=reason or "",
        )
        self._items.append(item)
        return item.item_id

    def add_items(self, items: list[dict]) -> int:
        """批量添加待确认项,返回添加数。

        Args:
            items: 每项为 dict,字段同 add_item 参数(缺字段用默认值)

        Returns:
            成功添加的数量(格式异常的项会被跳过,不中断批量)
        """
        count = 0
        for it in items or []:
            try:
                self.add_item(
                    task_id=it.get("task_id", ""),
                    question_num=it.get("question_num", ""),
                    stem=it.get("stem", ""),
                    options=it.get("options", []) or [],
                    garbled_answer=it.get("garbled_answer", ""),
                    suggested_answer=it.get("suggested_answer"),
                    confidence=float(it.get("confidence", 0.0)),
                    reason=it.get("reason", ""),
                )
                count += 1
            except (TypeError, ValueError):
                # 跳过格式异常的项,不中断批量
                continue
        return count

    def list_items(self, task_id: str | None = None) -> list[PendingItem]:
        """列出待确认项。

        Args:
            task_id: 指定任务 ID;None 列出全部

        Returns:
            PendingItem 列表
        """
        if task_id is None:
            return list(self._items)
        return [it for it in self._items if it.task_id == task_id]

    def confirm(self, item_id: str, answer: str) -> ExpertResult:
        """人工确认答案。

        Args:
            item_id: 待确认项 ID
            answer: 确认的答案

        Returns:
            ExpertResult(output=item_id, metadata={status})
        """
        target: PendingItem | None = None
        for it in self._items:
            if it.item_id == item_id:
                target = it
                break
        if target is None:
            return self._failure(f"pending item not found: {item_id}")
        if not answer or not str(answer).strip():
            return self._failure("answer must be non-empty")
        target.confirmed_answer = str(answer)
        target.status = "confirmed"
        print(f"[audit] pending item confirmed: id={item_id} answer={answer}")
        return self._success(item_id, status="confirmed")

    # ------------------------------------------------------------------
    # 行转换(共用)
    # ------------------------------------------------------------------

    def _row_of(self, item: PendingItem) -> list[Any]:
        """将 PendingItem 转为导出行(顺序与 _EXPORT_COLUMNS 一致)。"""
        return [
            _to_cell(item.question_num),
            _to_cell(item.stem),
            _to_cell(_get_option(item.options, 0)),
            _to_cell(_get_option(item.options, 1)),
            _to_cell(_get_option(item.options, 2)),
            _to_cell(_get_option(item.options, 3)),
            _to_cell(item.garbled_answer),
            _to_cell(item.suggested_answer),
            item.confidence,
            _to_cell(item.reason),
            _to_cell(item.status),
            _to_cell(item.confirmed_answer),
        ]

    # ------------------------------------------------------------------
    # 导出
    # ------------------------------------------------------------------

    def export_excel(
        self,
        output_path: str,
        task_id: str | None = None,
    ) -> ExpertResult:
        """导出为 xlsx。

        样式:表头加粗 + 浅蓝背景 + 居中;列宽自适应(中文按 2 计);冻结首行。

        Args:
            output_path: 输出 .xlsx 路径
            task_id: 仅导出指定任务的项;None 全部

        Returns:
            ExpertResult(output=output_path, rows=N)
        """
        err = self._validate_path(output_path, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ExpertError as e:
            return self._failure(str(e))

        items = self.list_items(task_id)
        wb = None
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Pending"

            headers = [c[0] for c in _EXPORT_COLUMNS]
            ws.append(headers)

            # 表头样式:加粗 + 浅蓝背景 + 居中
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # 数据行
            for item in items:
                ws.append(self._row_of(item))

            # 列宽自适应:遍历每列计算最大显示宽度(中文按 2 计)
            for col_idx in range(1, len(headers) + 1):
                # 表头宽度
                max_len = _display_width(str(headers[col_idx - 1]))
                # 数据行宽度
                for row in ws.iter_rows(
                    min_row=2, min_col=col_idx, max_col=col_idx, values_only=True
                ):
                    val = row[0] if row else None
                    if val is None:
                        continue
                    width = _display_width(str(val))
                    if width > max_len:
                        max_len = width
                # padding 2,下限 8,上限 60
                ws.column_dimensions[get_column_letter(col_idx)].width = min(
                    max(max_len + 2, 8), 60
                )

            # 冻结首行
            ws.freeze_panes = "A2"

            wb.save(output_path)
            return self._success(output_path, rows=len(items))
        except (OSError, PermissionError) as e:
            return self._failure(f"export_excel IO failed: {e}")
        except Exception as e:
            return self._failure(f"export_excel failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    _logger.debug('Unhandled exception', exc_info=True)

    def export_csv(
        self,
        output_path: str,
        task_id: str | None = None,
    ) -> ExpertResult:
        """导出为 CSV。

        编码 utf-8-sig(兼容 Excel 中文直接打开)。

        Args:
            output_path: 输出 .csv 路径
            task_id: 仅导出指定任务的项;None 全部

        Returns:
            ExpertResult(output=output_path, rows=N)
        """
        # CSV 不做扩展名校验(部分场景用 .txt),但仍需非空路径
        err = self._validate_string(output_path, "output_path")
        if err:
            return self._failure(err)

        items = self.list_items(task_id)
        try:
            # newline="" 避免 Windows 多余空行;utf-8-sig 兼容 Excel 中文
            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([c[0] for c in _EXPORT_COLUMNS])
                for item in items:
                    writer.writerow(self._row_of(item))
            return self._success(output_path, rows=len(items))
        except (OSError, PermissionError) as e:
            return self._failure(f"export_csv IO failed: {e}")
        except Exception as e:
            return self._failure(f"export_csv failed: {e}")

    def import_confirmed(self, xlsx_path: str) -> ExpertResult:
        """从 xlsx 导入人工确认的答案。

        读取"确认答案"列非空的行,返回 [{question_num, answer}] 列表,
        同时同步到内存中匹配的 PendingItem(按 question_num 匹配)。

        Args:
            xlsx_path: xlsx 文件路径(需含"题号"与"确认答案"列表头)

        Returns:
            ExpertResult(output=[{question_num, answer}, ...],
                         metadata={rows, synced})
        """
        err = self._validate_path(xlsx_path, must_exist=True, allowed_exts=("xlsx",))
        if err:
            return self._failure(err)

        try:
            self._require_lib("openpyxl")
            from openpyxl import load_workbook
        except ExpertError as e:
            return self._failure(str(e))

        wb = None
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = list(next(rows_iter))
            except StopIteration:
                return self._success(output=[], rows=0, synced=0)

            # 定位"题号"与"确认答案"列(按表头名匹配)
            qn_idx: int | None = None
            ans_idx: int | None = None
            for i, h in enumerate(headers):
                if h is None:
                    continue
                hs = str(h).strip()
                if hs == "题号":
                    qn_idx = i
                elif hs == "确认答案":
                    ans_idx = i
            if qn_idx is None or ans_idx is None:
                return self._failure("xlsx missing required columns: '题号' and '确认答案'")

            confirmed: list[dict] = []
            for row in rows_iter:
                if row is None:
                    continue
                qn = row[qn_idx] if qn_idx < len(row) else None
                ans = row[ans_idx] if ans_idx < len(row) else None
                if ans is None or str(ans).strip() == "":
                    continue
                confirmed.append(
                    {
                        "question_num": str(qn) if qn is not None else "",
                        "answer": str(ans),
                    }
                )

            # 同步到内存中的 items(按 question_num 匹配)
            qn_to_answer = {c["question_num"]: c["answer"] for c in confirmed}
            synced = 0
            for it in self._items:
                if it.question_num in qn_to_answer:
                    it.confirmed_answer = qn_to_answer[it.question_num]
                    it.status = "confirmed"
                    synced += 1

            return self._success(
                output=confirmed,
                rows=len(confirmed),
                synced=synced,
            )
        except (OSError, PermissionError) as e:
            return self._failure(f"import_confirmed IO failed: {e}")
        except Exception as e:
            return self._failure(f"import_confirmed failed: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    _logger.debug('Unhandled exception', exc_info=True)

    # ------------------------------------------------------------------
    # 统计
    # ------------------------------------------------------------------

    def summary(self) -> dict:
        """统计:总数/待确认/已确认/平均置信度。

        Returns:
            dict(total, pending, confirmed, avg_confidence)
        """
        total = len(self._items)
        pending = sum(1 for it in self._items if it.status == "pending")
        confirmed = sum(1 for it in self._items if it.status == "confirmed")
        avg_conf = sum(it.confidence for it in self._items) / total if total > 0 else 0.0
        return {
            "total": total,
            "pending": pending,
            "confirmed": confirmed,
            "avg_confidence": round(avg_conf, 4),
        }


# ---------------------------------------------------------------------------
# 辅助
# ---------------------------------------------------------------------------


def _display_width(s: str) -> int:
    """计算字符串显示宽度(中文字符按 2 计,ASCII 按 1 计)。

    用于列宽自适应估算。
    """
    if not s:
        return 0
    return sum(2 if ord(ch) > 127 else 1 for ch in s)
