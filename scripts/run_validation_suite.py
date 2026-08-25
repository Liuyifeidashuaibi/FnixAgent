# -*- coding: utf-8 -*-
"""真实任务验证跑批器 — 用真实 LLM 审判 Work 流水线（Forge 哲学：确定性判定，无 LLM 裁判）。

用法:
    python scripts/run_validation_suite.py --manifest bench/validation/tasks_work_v1.json \
        [--only W01,W02] [--limit N] [--timeout 240]

结果: 追加写入 bench_results/validation_runs.jsonl；结束打印汇总表。
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import tempfile
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))
os.chdir(ROOT)

RESULT_LOG = ROOT / "bench_results" / "validation_runs.jsonl"


def run_task(client, task: dict, timeout: int) -> dict:
    workspace = tempfile.mkdtemp(prefix=f"val_{task['id']}_")
    # 预置文件
    for rel, content in (task.get("pre_seed") or {}).items():
        p = Path(workspace) / rel
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content, encoding="utf-8")

    record = {
        "id": task["id"],
        "difficulty": task.get("difficulty", ""),
        "started": time.strftime("%H:%M:%S"),
        "workspace": workspace,
    }
    t0 = time.time()
    try:
        with client.stream(
            "POST",
            "/api/v1/work/stream",
            json={"user_input": task["prompt"], "workspace": workspace, "work_mode": "craft"},
            timeout=timeout,
        ) as resp:
            record["http"] = resp.status_code
            if resp.status_code != 200:
                record["ok"] = False
                record["error"] = f"http {resp.status_code}"
                return finish(record, t0)
            done_seen = False
            for line in resp.iter_lines():
                if not line.strip():
                    continue
                try:
                    obj = json.loads(line)
                except Exception:
                    continue
                ct = obj.get("chunk_type")
                if ct == "done":
                    done_seen = True
                    data = obj.get("content") or {}
                    record["trace_id"] = data.get("trace_id", "")
                elif ct == "error":
                    record["pipeline_error"] = str(obj.get("content"))[:300]
            record["done_seen"] = done_seen
    except Exception as exc:  # noqa: BLE001
        record["ok"] = False
        record["error"] = f"{type(exc).__name__}: {exc}"[:300]
        return finish(record, t0)

    record["checks"] = apply_checks(task.get("checks", []), workspace)
    passed = [c for c in record["checks"] if c["pass"]]
    record["ok"] = bool(record["checks"]) and len(passed) == len(record["checks"])
    return finish(record, t0)


def finish(record: dict, t0: float) -> dict:
    record["duration_s"] = round(time.time() - t0, 1)
    record.setdefault("ok", False)
    return record


def apply_checks(checks: list[dict], workspace: str) -> list[dict]:
    out = []
    for c in checks:
        ctype, path = c["type"], c.get("path", "")
        full = Path(workspace) / path
        result = {"type": ctype, "path": path, "pass": False}
        try:
            if not full.exists():
                out.append(result)
                continue
            if ctype == "file_exists":
                result["pass"] = True
            elif ctype in ("contains", "regex", "not_contains"):
                text = full.read_text(encoding="utf-8", errors="replace")
                needle = c.get("value", "")
                if ctype == "contains":
                    result["pass"] = needle in text
                elif ctype == "not_contains":
                    result["pass"] = needle not in text
                else:
                    import re as _re

                    result["pass"] = bool(_re.search(needle, text))
            elif ctype == "json_valid":
                json.loads(full.read_text(encoding="utf-8"))
                result["pass"] = True
            elif ctype == "json_array_min":
                data = json.loads(full.read_text(encoding="utf-8"))
                arr = data.get(c.get("key"), []) if isinstance(data, dict) else []
                result["pass"] = isinstance(arr, list) and len(arr) >= int(c.get("min", 1))
            elif ctype == "xlsx_exists":
                from openpyxl import load_workbook

                wb = load_workbook(str(full))
                result["pass"] = wb.sheetnames and len(wb.sheetnames) >= 1
            elif ctype == "docx_exists":
                from docx import Document

                doc = Document(str(full))
                result["pass"] = len(doc.paragraphs) >= 1
        except Exception as exc:  # noqa: BLE001 — 判定异常即不通过
            result["error"] = f"{type(exc).__name__}: {exc}"[:120]
        out.append(result)
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--manifest", required=True)
    ap.add_argument("--only", default="")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--timeout", type=int, default=240)
    args = ap.parse_args()

    manifest = json.loads(Path(args.manifest).read_text(encoding="utf-8"))
    tasks = manifest["tasks"]
    if args.only:
        wanted = {x.strip() for x in args.only.split(",")}
        tasks = [t for t in tasks if any(t["id"] == w or t["id"].startswith(w + "_") for w in wanted)]
    if args.limit:
        tasks = tasks[: args.limit]

    from fastapi.testclient import TestClient
    import fnixagent.main as m

    records = []
    with TestClient(m.app) as client:
        for i, task in enumerate(tasks, 1):
            print(f"[{i}/{len(tasks)}] {task['id']} ...", flush=True)
            rec = run_task(client, task, args.timeout)
            records.append(rec)
            mark = "PASS" if rec["ok"] else "FAIL"
            detail = ""
            if not rec["ok"]:
                bad = [c for c in rec.get("checks", []) if not c["pass"]]
                detail = " | " + ("; ".join(c["path"] + ":" + c.get("error", c["type"]) for c in bad[:2]) if bad else rec.get("error", "?"))
            print(f"    -> {mark} ({rec['duration_s']}s){detail}", flush=True)
            RESULT_LOG.parent.mkdir(parents=True, exist_ok=True)
            with open(RESULT_LOG, "a", encoding="utf-8") as f:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    ok_n = sum(1 for r in records if r["ok"])
    print("\n===== 汇总 =====")
    for r in records:
        print(f"  {'✅' if r['ok'] else '❌'} {r['id']:<26} {r['duration_s']:>6}s")
    print(f"\n通过率: {ok_n}/{len(records)} = {ok_n / max(len(records), 1):.0%}")


if __name__ == "__main__":
    main()
